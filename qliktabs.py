"""Script mínimo y limpio que funciona en tu entorno.

Flujo:
- abre la URL en Chrome
- escribe `username` carácter a carácter con Selenium
- si no aparece, envía texto a nivel sistema (Windows)
- usa `pywinauto` para enviar Tab+Enter y activar el botón
- mantiene la ventana abierta unos segundos para inspección
"""
from __future__ import annotations

import logging
import time
import platform
import ctypes
import urllib.parse
from ctypes import wintypes
import os
import glob
import json
from pathlib import Path
from datetime import datetime, timedelta
import os as _os
import re

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

LOG = logging.getLogger(__name__)

tiempo=30
corto_tiempo=2

def setup_driver() -> webdriver.Chrome:
	opts = Options()
	opts.add_argument("--no-sandbox")
	opts.add_argument("--disable-dev-shm-usage")
	service = Service(ChromeDriverManager().install())
	driver = webdriver.Chrome(service=service, options=opts)
	driver.maximize_window()
	return driver


def type_like_keyboard(driver: webdriver.Chrome, text: str, delay: float = 0.08, click_first: bool = True) -> bool:
	try:
		if click_first:
			try:
				active = driver.switch_to.active_element
				if not active or active.tag_name.lower() == 'html':
					try:
						driver.find_element(By.TAG_NAME, 'body').click()
					except Exception:
						pass
					time.sleep(2)
			except Exception:
				time.sleep(2)

		active = driver.switch_to.active_element
		if not active:
			return False

		sent = 0
		for ch in text:
			try:
				active.send_keys(ch)
				sent += 1
				time.sleep(delay)
			except Exception:
				continue
		LOG.info("Typist: enviados %d/%d caracteres", sent, len(text))
		return sent > 0
	except Exception:
		LOG.exception("Error en type_like_keyboard")
		return False


def _send_text_windows(text: str, delay: float = 0.08) -> None:
	if platform.system() != "Windows":
		return
	user32 = ctypes.WinDLL('user32', use_last_error=True)
	VkKeyScanW = user32.VkKeyScanW
	keybd_event = user32.keybd_event
	VK_SHIFT = 0x10
	KEYEVENTF_KEYUP = 0x0002
	for ch in text:
		vks = VkKeyScanW(ord(ch))
		if vks == -1:
			continue
		vk = vks & 0xFF
		shift_state = (vks >> 8) & 0xFF
		if shift_state & 1:
			keybd_event(VK_SHIFT, 0, 0, 0)
		keybd_event(vk, 0, 0, 0)
		time.sleep(0.01)
		keybd_event(vk, 0, KEYEVENTF_KEYUP, 0)
		if shift_state & 1:
			keybd_event(VK_SHIFT, 0, KEYEVENTF_KEYUP, 0)
		time.sleep(delay)


def send_text_via_system(driver: webdriver.Chrome, text: str, delay: float = 0.08) -> bool:
	try:
		try:
			driver.execute_script("window.focus();")
		except Exception:
			pass
		time.sleep(0.2)
		if platform.system() == "Windows":
			_send_text_windows(text, delay=delay)
			LOG.info("Envío de texto por sistema completado (Windows)")
			return True
		return False
	except Exception:
		LOG.exception("Fallo en send_text_via_system")
		return False


def send_keys_via_pywinauto(keys: str, url_substring: str | None = None) -> bool:
	try:
		from pywinauto import Desktop, keyboard
	except Exception:
		LOG.debug('pywinauto no instalado')
		return False
	try:
		desktop = Desktop(backend="uia")
		windows = desktop.windows()
		target = None
		us = url_substring.lower() if url_substring else None
		for w in windows:
			try:
				title = (w.window_text() or '').lower()
				if not title:
					continue
				if us and us in title:
					target = w
					break
				if not us and ('chrome' in title or 'qlik' in title):
					target = w
					break
			except Exception:
				continue
		if not target:
			for w in windows:
				try:
					title = (w.window_text() or '').lower()
					if 'qlik' in title or 'chrome' in title:
						target = w
						break
				except Exception:
					continue
		if not target:
			return False
		target.set_focus()
		time.sleep(0.12)
		keyboard.send_keys(keys)
		time.sleep(0.12)
		LOG.info('send_keys_via_pywinauto: keys enviados')
		return True
	except Exception:
		LOG.debug('send_keys_via_pywinauto: fallo', exc_info=True)
		return False


def _send_enter_windows() -> bool:
	if platform.system() != 'Windows':
		return False
	try:
		user32 = ctypes.WinDLL('user32', use_last_error=True)
		keybd_event = user32.keybd_event
		KEYEVENTF_KEYUP = 0x0002
		VK_RETURN = 0x0D
		keybd_event(VK_RETURN, 0, 0, 0)
		time.sleep(1)
		keybd_event(VK_RETURN, 0, KEYEVENTF_KEYUP, 0)
		time.sleep(1)
		LOG.info('_send_enter_windows: Enter enviado')
		return True
	except Exception:
		LOG.debug('_send_enter_windows: fallo', exc_info=True)
		return False


def focus_on_selector(driver: webdriver.Chrome, selector: str, timeout: float = 3.0) -> bool:
	"""Intentar enfocar/posicionar el elemento identificado por `selector`.

	Estrategia:
	- buscar con Selenium y hacer scrollIntoView
	- intentar click directo
	- fallback a ActionChains move_to_element + click
	- usar execute_script focus() si sigue sin foco
	- si todo falla, devolver False
	"""
	try:
		end = time.time() + float(timeout)
		while time.time() < end:
			try:
				el = driver.find_element(By.CSS_SELECTOR, selector)
			except Exception:
				el = None
			if not el:
				time.sleep(1)
				continue

			try:
				# asegurar que está visible en pantalla
				driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'nearest'});", el)
			except Exception:
				pass

			try:
				el.click()
				LOG.info("focus_on_selector: click directo en %s", selector)
			except Exception:
				try:
					ActionChains(driver).move_to_element(el).click().perform()
					LOG.info("focus_on_selector: click via ActionChains en %s", selector)
				except Exception:
					LOG.debug("focus_on_selector: el.click/ActionChains falló, intentando focus()")
					try:
						driver.execute_script("arguments[0].focus();", el)
					except Exception:
						pass

			try:
				# intento simple para forzar foco en el elemento
				el.send_keys("")
			except Exception:
				pass

			# comprobar si el elemento es el activo
			try:
				active = driver.switch_to.active_element
				if active is not None:
					try:
						if el == active or (active.get_attribute('outerHTML') and el.get_attribute('outerHTML') in active.get_attribute('outerHTML')):
							LOG.info("focus_on_selector: elemento activo detectado para %s", selector)
							return True
					except Exception:
						# si no podemos comparar, aún aceptamos si el elemento parece visible
						return True
				else:
					return True
			except Exception:
				return True

		LOG.debug("focus_on_selector: no se encontró o no se pudo enfocar %s", selector)
		return False
	except Exception:
		LOG.exception("focus_on_selector: excepción inesperada")
		return False


def hover_on_selector(driver: webdriver.Chrome, selector: str, timeout: float = 5.0) -> bool:
	"""Simular hover (mover el cursor) sobre el elemento y despachar eventos mouseover/mouseenter.

	Esto usa ActionChains para mover el cursor virtual y también despacha events via JS
	porque algunas UIs reaccionan sólo a eventos de mouse.
	"""
	try:
		end = time.time() + float(timeout)
		while time.time() < end:
			try:
				el = driver.find_element(By.CSS_SELECTOR, selector)
			except Exception:
				el = None
			if not el:
				time.sleep(15)
				continue

			try:
				driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'nearest'});", el)
			except Exception:
				pass

			try:
				ActionChains(driver).move_to_element(el).perform()
				LOG.info("hover_on_selector: move_to_element realizado para %s", selector)
			except Exception:
				LOG.debug("hover_on_selector: ActionChains falló", exc_info=True)

			try:
				js = (
					"var e = new MouseEvent('mouseover', {bubbles:true, cancelable:true});"
					"arguments[0].dispatchEvent(e);"
					"var e2 = new MouseEvent('mouseenter', {bubbles:true, cancelable:true});"
					"arguments[0].dispatchEvent(e2);"
				)
				driver.execute_script(js, el)
				LOG.info("hover_on_selector: eventos mouseover/mouseenter despachados para %s", selector)
			except Exception:
				LOG.debug("hover_on_selector: dispatch JS falló", exc_info=True)

			time.sleep(0.2)
			return True

		LOG.debug("hover_on_selector: no se encontró %s", selector)
		return False
	except Exception:
		LOG.exception("hover_on_selector: excepción inesperada")
		return False


def hover_on_xpath(driver: webdriver.Chrome, xpath: str, timeout: float = 5.0) -> bool:
	"""Similar a `hover_on_selector` pero busca por XPath en lugar de CSS selector."""
	try:
		end = time.time() + float(timeout)
		while time.time() < end:
			try:
				el = driver.find_element(By.XPATH, xpath)
			except Exception:
				el = None
			if not el:
				time.sleep(1)
				continue

			try:
				driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'nearest'});", el)
			except Exception:
				pass

			try:
				ActionChains(driver).move_to_element(el).perform()
				LOG.info("hover_on_xpath: move_to_element realizado para %s", xpath)
			except Exception:
				LOG.debug("hover_on_xpath: ActionChains falló", exc_info=True)

			try:
				js = (
					"var e = new MouseEvent('mouseover', {bubbles:true, cancelable:true});"
					"arguments[0].dispatchEvent(e);"
					"var e2 = new MouseEvent('mouseenter', {bubbles:true, cancelable:true});"
					"arguments[0].dispatchEvent(e2);"
				)
				driver.execute_script(js, el)
				LOG.info("hover_on_xpath: eventos mouseover/mouseenter despachados para %s", xpath)
			except Exception:
				LOG.debug("hover_on_xpath: dispatch JS falló", exc_info=True)

			time.sleep(corto_tiempo)
			return True

		LOG.debug("hover_on_xpath: no se encontró %s", xpath)
		return False
	except Exception:
		LOG.exception("hover_on_xpath: excepción inesperada")
		return False


def click_button_by_selector(driver: webdriver.Chrome, selector: str, timeout: float = 5.0) -> bool:
	"""Buscar y clicar un botón identificado por `selector`.

	Intentos: encontrar, scrollIntoView, click directo, ActionChains click, o click vía JS.
	"""
	try:
		end = time.time() + float(timeout)
		while time.time() < end:
			try:
				btn = driver.find_element(By.CSS_SELECTOR, selector)
			except Exception:
				btn = None
			if not btn:
				time.sleep(0.25)
				continue

			try:
				driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'nearest'});", btn)
			except Exception:
				pass

			try:
				btn.click()
				LOG.info("click_button_by_selector: click directo en %s", selector)
				return True
			except Exception:
				try:
					ActionChains(driver).move_to_element(btn).click().perform()
					LOG.info("click_button_by_selector: click via ActionChains en %s", selector)
					return True
				except Exception:
					try:
						driver.execute_script("arguments[0].click();", btn)
						LOG.info("click_button_by_selector: click via JS en %s", selector)
						return True
					except Exception:
						LOG.debug("click_button_by_selector: intento de click falló para %s", selector, exc_info=True)
			time.sleep(0.2)
		LOG.debug("click_button_by_selector: no se encontró %s", selector)
		return False
	except Exception:
		LOG.exception("click_button_by_selector: excepción inesperada")
		return False


def click_button_by_xpath(driver: webdriver.Chrome, xpath: str, timeout: float = 5.0) -> bool:
	"""Buscar y clicar un botón identificado por `xpath`.

	Intenta: encontrar, scrollIntoView, click directo, ActionChains click, o click vía JS.
	"""
	try:
		end = time.time() + float(timeout)
		while time.time() < end:
			try:
				btn = driver.find_element(By.XPATH, xpath)
			except Exception:
				btn = None
			if not btn:
				time.sleep(0.25)
				continue

			try:
				driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'nearest'});", btn)
			except Exception:
				pass

			try:
				btn.click()
				LOG.info("click_button_by_xpath: click directo en %s", xpath)
				return True
			except Exception:
				try:
					ActionChains(driver).move_to_element(btn).click().perform()
					LOG.info("click_button_by_xpath: click via ActionChains en %s", xpath)
					return True
				except Exception:
					try:
						driver.execute_script("arguments[0].click();", btn)
						LOG.info("click_button_by_xpath: click via JS en %s", xpath)
						return True
					except Exception:
						LOG.debug("click_button_by_xpath: intento de click falló para %s", xpath, exc_info=True)
			time.sleep(0.2)
		LOG.debug("click_button_by_xpath: no se encontró %s", xpath)
		return False
	except Exception:
		LOG.exception("click_button_by_xpath: excepción inesperada")
		return False


def click_export_url(driver: webdriver.Chrome, selector: str = 'a.export-url', timeout: float = 10.0) -> bool:
	"""Esperar y clicar el enlace de exportación visible con clase `export-url`.

	El href puede cambiar dinámicamente; por eso buscamos el ancla por clase y clickeamos.
	"""
	try:
		end = time.time() + float(timeout)
		while time.time() < end:
			try:
				a = driver.find_element(By.CSS_SELECTOR, selector)
			except Exception:
				a = None
			if not a:
				time.sleep(0.5)
				continue

			try:
				href = a.get_attribute('href') or a.get_attribute('ng-href')
			except Exception:
				href = None

			try:
				driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", a)
			except Exception:
				pass

			try:
				# Preferimos dispatch JS porque abre en nueva pestaña si target=_blank
				driver.execute_script("arguments[0].click();", a)
				LOG.info("click_export_url: clicado enlace %s (href=%s)", selector, href)
				return True
			except Exception:
				try:
					a.click()
					LOG.info("click_export_url: clicado elemento via click() %s (href=%s)", selector, href)
					return True
				except Exception:
					LOG.debug("click_export_url: intento de click falló para %s", selector, exc_info=True)

			time.sleep(0.2)
		LOG.debug("click_export_url: no se encontró %s dentro del timeout", selector)
		return False
	except Exception:
		LOG.exception("click_export_url: excepción inesperada")
		return False


def click_export_link_with_fallback(driver: webdriver.Chrome, timeout: float = 6.0) -> bool:
	"""Intentar localizar y clicar el enlace de descarga de export directamente.

	Estrategia:
	- intentar la clase conocida `a.export-url`
	- buscar anchors con href que contengan '.xlsx'
	- buscar anchors cuyo texto contenga 'export' o 'exportar' (case-insensitive)
	- click via JS o ActionChains
	"""
	try:
		# intento rápido con la clase que ya usamos
		if click_export_url(driver, selector='a.export-url', timeout=2.0):
			return True

		end = time.time() + float(timeout)
		xpaths = [
			"//a[contains(translate(@href,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'.xlsx') ]",
			"//a[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'export') ]",
			"//a[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'exportar') ]",
		]

		while time.time() < end:
			for xp in xpaths:
				try:
					els = driver.find_elements(By.XPATH, xp)
				except Exception:
					els = []
				if not els:
					continue
				for el in els:
					try:
						if not el.is_displayed():
							continue
					except Exception:
						pass
					try:
						driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
					except Exception:
						pass
					try:
						driver.execute_script("arguments[0].click();", el)
						LOG.info("click_export_link_with_fallback: clicado anchor por XPath: %s", xp)
						return True
					except Exception:
						try:
							ActionChains(driver).move_to_element(el).click().perform()
							LOG.info("click_export_link_with_fallback: clicado anchor via ActionChains por XPath: %s", xp)
							return True
						except Exception:
							LOG.debug("click_export_link_with_fallback: fallo al clicar anchor por XPath: %s", xp, exc_info=True)
			time.sleep(0.3)

		LOG.debug('click_export_link_with_fallback: no se encontró anchor de descarga dentro del timeout')
		return False
	except Exception:
		LOG.exception('click_export_link_with_fallback: excepción inesperada')
		return False


def bring_browser_to_front(driver: webdriver.Chrome, url_substring: str | None = None) -> bool:
	"""Intentar traer la ventana del navegador al frente.

	Estrategia (no intrusiva):
	- ejecutar `window.focus()` via JS
	- intentar usar `pywinauto` para localizar la ventana por título y `set_focus()` (Windows)
	- devolver True si alguno de los métodos parece haber tenido éxito
	"""
	try:
		try:
			driver.execute_script("window.focus();")
		except Exception:
			pass

		# Intentar pywinauto (solo si está instalado)
		try:
			from pywinauto import Desktop
		except Exception:
			LOG.debug('bring_browser_to_front: pywinauto no disponible')
			return False

		try:
			desktop = Desktop(backend="uia")
			windows = desktop.windows()
			target = None
			us = url_substring.lower() if url_substring else None
			for w in windows:
				try:
					title = (w.window_text() or '').lower()
					if not title:
						continue
					if us and us in title:
						target = w
						break
					if not us and ('chrome' in title or 'qlik' in title):
						target = w
						break
				except Exception:
					continue

			if target:
				try:
					target.set_focus()
					LOG.info('bring_browser_to_front: foco restaurado con pywinauto (ventana: %s)', target.window_text())
					return True
				except Exception:
					LOG.debug('bring_browser_to_front: fallo al set_focus()', exc_info=True)
					return False
		except Exception:
			LOG.debug('bring_browser_to_front: fallo buscando ventanas con pywinauto', exc_info=True)
			return False

		return False
	except Exception:
		LOG.debug('bring_browser_to_front: excepción inesperada', exc_info=True)
		return False


def find_latest_downloaded_file(directory: str, pattern: str = '*.xlsx', since_ts: float | None = None, timeout: float = 30.0) -> str | None:
	"""Buscar el fichero más reciente que coincida con pattern en `directory`.

	- `since_ts`: si se proporciona, devolver sólo archivos con mtime >= since_ts - 2s
	- devuelve la ruta absoluta o None si no se encuentra en el timeout
	"""
	end = time.time() + float(timeout)
	d = Path(directory).expanduser()
	if not d.exists():
		LOG.debug("find_latest_downloaded_file: directorio no existe %s", d)
		return None

	while time.time() < end:
		files = list(d.glob(pattern))
		if not files:
			time.sleep(0.5)
			continue
		# escoger el más reciente
		files_sorted = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)
		candidate = files_sorted[0]
		try:
			mtime = candidate.stat().st_mtime
		except Exception:
			time.sleep(0.5)
			continue
		if since_ts is None or mtime >= (since_ts - 2.0):
			return str(candidate)
		time.sleep(0.5)
	return None


def extract_excel_contents(path: str) -> dict | None:
	"""Extraer contenido del Excel en `path`.

	Retorna un dict {sheet_name: [row_dicts]}. Usa pandas si está disponible, else openpyxl.
	"""
	def format_number_es(value: float, decimals: int) -> str:
		# Formatear número con separador de miles '.' y decimal ','
		try:
			fmt = f"{abs(value):,.{decimals}f}"
			# swap comma and dot to European format
			tmp = fmt.replace(',', 'X')
			tmp = tmp.replace('.', ',')
			tmp = tmp.replace('X', '.')
			return tmp
		except Exception:
			return str(value)

	def format_cell_display(cell) -> str:
		try:
			val = cell.value
			nf = (cell.number_format or '').lower()
			if val is None:
				return ''
			# Dates
			if hasattr(cell, 'is_date') and cell.is_date:
				try:
					if isinstance(val, datetime):
						return val.isoformat()
					return str(val)
				except Exception:
					return str(val)

			# Numeric formatting
			if isinstance(val, (int, float)):
				# percentage
				if '%' in nf:
					# decidir decimales por la presencia de '0.0' en el formato
					decimals = 1 if '0.0' in nf or '0,0' in nf else 0
					perc = val * 100
					s = format_number_es(perc, decimals)
					return ("-" + s + "%") if perc < 0 else (s + "%")

				# currency
				if '$' in nf or '€' in nf or '¤' in nf:
					# intentar deducir decimales
					decimals = 0
					if '0.00' in nf or '0,00' in nf:
						decimals = 2
					s = format_number_es(val, decimals)
					sign = '-' if val < 0 else ''
					symbol = '$' if '$' in nf else ('€' if '€' in nf else '')
					# Devolver número limpio (sin signo de moneda ni separadores)
					cleaned = re.sub(r'[^0-9\-]', '', sign + s)
					return cleaned

				# default numeric
				# si es prácticamente entero, no mostrar decimales
				if abs(val - round(val)) < 0.005:
					s = format_number_es(round(val), 0)
					return s
				else:
					s = format_number_es(val, 2)
					return s

			# strings
			return str(val)
		except Exception:
			try:
				return str(cell.value)
			except Exception:
				return ''

	try:
		# Preferir openpyxl para conservar formatos mostrados
		try:
			from openpyxl import load_workbook
			wb = load_workbook(path, data_only=True)
			out = {}
			for sheet in wb.sheetnames:
				ws = wb[sheet]
				rows = list(ws.rows)
				if not rows:
					out[sheet] = []
					continue
				headers = [ (c.value if c.value is not None else f'col{i}') for i, c in enumerate(rows[0], start=1) ]
				data = []
				for r in rows[1:]:
					rowd = {}
					for h, cell in zip(headers, r):
						rowd[str(h)] = format_cell_display(cell)
					data.append(rowd)
				out[sheet] = data
			return out
		except Exception:
			LOG.debug('openpyxl no disponible o falló, intentando pandas', exc_info=True)
			try:
				import pandas as pd
				xls = pd.ExcelFile(path)
				result = {}
				for sheet in xls.sheet_names:
					df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
					df = df.fillna('')
					# Normalizar valores que contengan '$' eliminando símbolos y separadores
					records = df.to_dict(orient='records')
					for row in records:
						for k, v in list(row.items()):
							if isinstance(v, str) and '$' in v:
								row[k] = re.sub(r'[^0-9\-]', '', v)
					result[sheet] = records
				return result
			except Exception:
				LOG.exception('extract_excel_contents: No se pudo leer con pandas')
				return None
	except Exception:
		LOG.exception('extract_excel_contents: excepción inesperada')
		return None


def upload_to_google_sheets(extracted: dict, spreadsheet_id: str, credentials_json_path: str, clear: bool = True, target_sheet: str | None = 'Sheet2') -> bool:
	"""Subir `extracted` (dict sheet -> list[dict]) a Google Sheets.

	- `extracted`: dict devuelto por `extract_excel_contents`.
	- `spreadsheet_id`: id del spreadsheet (la parte larga de la URL /spreadsheets/d/<id>/... ).
	- `credentials_json_path`: ruta al JSON de la cuenta de servicio (service account).
	- `clear`: si True se borra la worksheet antes de escribir.

	Requiere: `gspread` y `google-auth` (google-auth). Si no están instalados, la función registra y devuelve False.
	"""
	try:
		try:
			import gspread
			from google.oauth2 import service_account
		except Exception:
			LOG.exception('upload_to_google_sheets: faltan dependencias (gspread/google-auth)')
			return False

		scopes = [
			'https://www.googleapis.com/auth/spreadsheets',
			'https://www.googleapis.com/auth/drive'
		]
		creds = service_account.Credentials.from_service_account_file(credentials_json_path, scopes=scopes)
		client = gspread.authorize(creds)

		sh = client.open_by_key(spreadsheet_id)

		# helper: sanitize cell values before uploading to Sheets
		def _sanitize_cell_value(v: object):
			"""Sanitize a cell value and coerce numeric-like strings to int/float when appropriate.

			Returns either an int, float, or string (or empty string for None).
			This removes leading apostrophes (ASCII and typographic), trims whitespace,
			and attempts to parse European/US formatted numbers (thousands separators
			and decimal separators) into numeric Python types so gspread writes numeric
			cells into Google Sheets.
			"""
			try:
				if v is None:
					return ''
				s = str(v).strip()

				# remove leading common quotes/apostrophes that force text in Sheets
				while s and s[0] in ("'", "\u2019", "\u2018", "`"):
					s = s[1:].lstrip()

				# after cleaning, if empty -> return empty string
				if s == '':
					return ''

				# helper: try integer parse by stripping non-digits (keep minus)
				def _try_int(x: str):
					cleaned = re.sub(r'[^0-9\-]', '', x)
					if cleaned and re.match(r'^-?\d+$', cleaned):
						try:
							return int(cleaned)
						except Exception:
							return None
					return None

				# helper: try float parse handling thousands and decimal separators
				def _try_float(x: str):
					t = x.replace(' ', '')
					# If contains both '.' and ',', decide decimal separator by last occurrence
					if '.' in t and ',' in t:
						if t.rfind(',') > t.rfind('.'):
							# comma likely decimal, dots thousands
							t2 = t.replace('.', '').replace(',', '.')
						else:
							# dot likely decimal, commas thousands
							t2 = t.replace(',', '')
					else:
						# only comma present -> treat comma as decimal
						if ',' in t and '.' not in t:
							t2 = t.replace('.', '').replace(',', '.')
						else:
							# treat commas as thousands separators
							t2 = t.replace(',', '')

					# remove any non-numeric/decimal/minus characters
					t2 = re.sub(r'[^0-9\.\-]', '', t2)
					if re.match(r'^-?\d+(?:\.\d+)?$', t2):
						try:
							return float(t2)
						except Exception:
							return None
					return None

				# Try int first (preferred), then float
				intval = _try_int(s)
				if intval is not None:
					return intval
				fl = _try_float(s)
				if fl is not None:
					# if it's effectively an integer (e.g. 123.0) return int
					if abs(fl - round(fl)) < 1e-9:
						return int(round(fl))
					return fl

				# fallback: return cleaned string
				return s
			except Exception:
				try:
					return str(v)
				except Exception:
					return ''

		# Simple sanitizer that ONLY strips a leading apostrophe/quote and left whitespace
		def _strip_leading_apostrophe(v: object):
			try:
				if v is None:
					return ''
				s = str(v)
				s = s.lstrip()
				while s and s[0] in ("'", "\u2019", "\u2018", "`"):
					s = s[1:].lstrip()
				return s
			except Exception:
				try:
					return str(v)
				except Exception:
					return ''

		# Identity sanitizer (leave the value as-is, used to avoid touching Sheet1)
		def _identity_sanitize(v: object):
			if v is None:
				return ''
			return v

		# Special sanitizer for Sheet1 column C: preserve the displayed formatting
		# (commas/dots) but DROP the last TWO characters of the displayed string.
		# Example: '407,918,004' -> '407,918,0' ; '24,774,107,615' -> '24,774,107,6'
		def _strip_dots_and_drop_decimals(v: object):
			try:
				if v is None:
					return ''
				s = str(v).strip()
				# strip leading common quotes/apostrophes that force text in Sheets
				while s and s[0] in ("'", "\u2019", "\u2018", "`"):
					s = s[1:].lstrip()
				# If the displayed string is short, return empty
				if len(s) <= 2:
					return ''
				# Remove the last two characters but preserve the rest (including separators)
				out = s[:-2].rstrip()
				return out
			except Exception:
				try:
					return str(v)
				except Exception:
					return ''

		# Si se indicó un target_sheet concreto, escribir la PRIMERA hoja extraída en esa hoja
		if target_sheet:
			try:
				first_sheet = next(iter(extracted.keys()))
				rows = extracted.get(first_sheet, [])
				safe_name = str(target_sheet)[:100]
				try:
					ws = sh.worksheet(safe_name)
				except Exception:
					ws = sh.add_worksheet(title=safe_name, rows=max(100, len(rows) + 5), cols=max(10, len(rows[0]) if rows else 10))

				# Preserve existing header row if present; otherwise use extracted headers.
				try:
					existing_headers = ws.row_values(1)
				except Exception:
					existing_headers = []

				if not existing_headers:
					# No header present in sheet -> derive from extracted data (if any)
					headers = list(rows[0].keys()) if rows else []
					existing_headers = headers

				# Decide header for column A (fecha) and data headers for B..
				def _norm(s: str) -> str:
					return re.sub(r'\s+', ' ', str(s).strip().lower())

				# Detect if sheet already has a fecha/date column in A1
				header_a = 'fecha'
				data_headers = list(existing_headers)
				if data_headers:
					first_norm = _norm(data_headers[0])
					if 'fecha' in first_norm or 'date' in first_norm:
						header_a = existing_headers[0]
						data_headers = existing_headers[1:]

				# Clear entire sheet and write back the header row starting at B1 (A1 will hold header_a)
				date_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
				# If clear requested, remove only data rows (row 2 and below) to preserve row 1
				if clear:
					try:
						# clear all rows from 2 to the current row_count
						try:
							rc = getattr(ws, 'row_count', None)
							if rc and isinstance(rc, int) and rc > 1:
								ws.batch_clear([f"2:{rc}"])
							else:
								# fallback: clear a large range of rows to be safe
								ws.batch_clear(["2:1000"])
						except Exception:
							# final fallback
							ws.batch_clear(["2:1000"])
					except Exception:
						pass

					# Additionally, ensure columns D and E (from row 2 downward) are cleared
					# but only for the Sheet named 'Sheet1' to avoid touching other tabs.
					try:
						try:
							sn_local = str(safe_name).strip().lower() if safe_name else ''
							if sn_local == 'sheet1':
								rc_cols = getattr(ws, 'row_count', None) or 1000
								if isinstance(rc_cols, int) and rc_cols > 1:
									ws.batch_clear([f"D2:D{rc_cols}", f"E2:E{rc_cols}"])
								else:
									ws.batch_clear(["D2:D1000", "E2:E1000"])
						except Exception:
							# best-effort fallback when attempting to clear
							try:
								if sn_local == 'sheet1':
									ws.batch_clear(["D2:D1000", "E2:E1000"])
							except Exception:
								pass
					except Exception:
						pass

				try:
					# Write header row only if the sheet has no header yet and
					# do NOT touch row 1 for Sheet1 or Sheet2 (preserve existing header)
					sn = str(safe_name).strip().lower() if safe_name else ''
					header_row = [header_a] + data_headers
					if not existing_headers and sn not in ('sheet1', 'sheet2'):
						try:
							ws.update('A1', [header_row])
						except Exception:
							LOG.debug('upload_to_google_sheets: fallo al escribir encabezado en %s', safe_name, exc_info=True)
				except Exception:
					LOG.debug('upload_to_google_sheets: fallo al preparar encabezado en %s', safe_name, exc_info=True)

				# Now write data rows (without headers) starting at A2 mapped to data_headers order.
				# Map data header names to extracted row keys using a tolerant normalization.
				if rows and data_headers is not None:
					try:
						extracted_keys = list(rows[0].keys()) if rows else []
						norm_map = { _norm(k): k for k in extracted_keys }

						# detect which extracted key is likely the per-row date
						date_key = None
						for ek in extracted_keys:
							enk = _norm(ek)
							if 'fecha' in enk or 'date' in enk or 'dia' in enk:
								date_key = ek
								break

						mapped_keys = []
						for h in data_headers:
							nh = _norm(h)
							mapped = norm_map.get(nh)
							if not mapped:
								nh_simple = re.sub(r'[^0-9a-z]', '', nh)
								for ek_norm, ek in norm_map.items():
									if re.sub(r'[^0-9a-z]', '', ek_norm) == nh_simple:
										mapped = ek
										break
							if not mapped:
								for ek_norm, ek in norm_map.items():
									if nh in ek_norm or ek_norm in nh:
										mapped = ek
										break
							mapped_keys.append(mapped)

							table_data = []
							# choose sanitizer per target sheet: only Sheet2 gets apostrophe-strip, Sheet1 and others are left untouched
							sn = str(safe_name).strip().lower() if safe_name else ''
							if sn == 'sheet2':
								base_sanitizer = _strip_leading_apostrophe
							else:
								base_sanitizer = _identity_sanitize

							# columns to coerce to numeric (sheet columns C,D,E,G,I,K) -> zero-based column indices
							numeric_col_letters = ['C', 'D', 'E', 'G', 'I', 'K']
							# convert letters to zero-based column numbers
							def _col_letter_to_index(letter: str) -> int:
								return ord(letter.upper()) - ord('A')
							numeric_col_indexes = {_col_letter_to_index(l) for l in numeric_col_letters}

							# data_headers correspond to sheet columns starting at B (col_index 1)
							# rowvals index -> sheet column index = index + 1
							numeric_rowvals_indexes = {ci - 1 for ci in numeric_col_indexes if ci >= 1}

							for r in rows:
								# decide per-row date: prefer extracted date_key, else use execution date
								if date_key:
									pv = r.get(date_key, '')
									per_row_date = str(pv) if pv not in (None, '') else date_str
								else:
									per_row_date = date_str

								rowvals = []
								for idx, mk in enumerate(mapped_keys):
									if mk:
										# raw value from extraction (do not pre-strip here for Sheet1 col C)
										rawv = r.get(mk, '')

										# Special-case: Sheet1 column C (data_headers start at B -> idx==1)
										# apply the drop-last-two sanitizer and DO NOT coerce/format it.
										if sn == 'sheet1' and idx == 1:
											try:
												vout = _strip_dots_and_drop_decimals(rawv)
											except Exception:
												vout = ''
											rowvals.append(vout)
											continue

										# Otherwise apply base sanitizer (may be identity or apostrophe-strip)
										try:
											raw = base_sanitizer(rawv)
										except Exception:
											raw = rawv if rawv is not None else ''

										# if this column should be numeric for Sheet2, attempt conversion
										if sn == 'sheet2' and idx in numeric_rowvals_indexes:
											try:
												coerced = _sanitize_cell_value(raw)
												rowvals.append(coerced)
											except Exception:
												rowvals.append(raw)
										else:
											rowvals.append(raw)
									else:
										rowvals.append('')
								# prefix with per-row date in column A
								table_data.append([per_row_date] + rowvals)

						try:
							# Clear known ARRAYFORMULA spill ranges that may block the formula from
							# expanding. Some sheets place an ARRAYFORMULA in E1 that spills into E2:E.
							# If those cells contain leftover data, the array won't expand and Sheets
							# raises "No se amplió el resultado del array porque reemplazaría los datos...".
							try:
								rc = getattr(ws, 'row_count', None) or 1000
								ws.batch_clear([f"E2:E{rc}"])
							except Exception:
								# best-effort fallback
								try:
									ws.batch_clear(["E2:E1000"])
								except Exception:
									pass

							# Pre-write: ensure D2:E... are cleared to avoid ARRAYFORMULA or residual data (Sheet1 only)
							sn_check = str(safe_name).strip().lower() if safe_name else ''
							if sn_check == 'sheet1':
								try:
									try:
										rc_pre = getattr(ws, 'row_count', None) or (1 + len(table_data))
										if isinstance(rc_pre, int) and rc_pre > 1:
											ws.batch_clear([f"D2:D{rc_pre}", f"E2:E{rc_pre}"])
										else:
											ws.batch_clear(["D2:D1000", "E2:E1000"])
									except Exception:
										ws.batch_clear(["D2:D1000", "E2:E1000"])
								except Exception:
									pass

							ws.update('A2', table_data)
							# Post-write: asegurar limpieza específica en D2:E... tras escribir los datos (Sheet1 only)
							if sn_check == 'sheet1':
								try:
									try:
										rc_after = getattr(ws, 'row_count', None) or (1 + len(table_data))
										if isinstance(rc_after, int) and rc_after > 1:
											ws.batch_clear([f"D2:D{rc_after}", f"E2:E{rc_after}"])
										else:
											ws.batch_clear(["D2:D1000", "E2:E1000"])
									except Exception:
										ws.batch_clear(["D2:D1000", "E2:E1000"])
								except Exception:
									pass
						except Exception:
							LOG.exception('upload_to_google_sheets: fallo al escribir datos en %s', safe_name)
						# Apply number formatting: Sheet2 columns C,D,E,G,I,K
						try:
							if sn == 'sheet2':
								# obtain sheetId for API requests
								try:
									sheet_id = int(ws._properties.get('sheetId'))
								except Exception:
									sheet_id = None
								if sheet_id is not None:
									# For Sheet2: columns C,D,E,G,I,K (zero-based: 2,3,4,6,8,10)
									# For Sheet1: column C only (zero-based: 2)
									if sn == 'sheet2':
										cols = [2, 3, 4, 6, 8, 10]
									else:  # sheet1
										cols = [2]
									requests = []
									for c in cols:
										requests.append({
											'repeatCell': {
												'range': {
													'sheetId': sheet_id,
													'startRowIndex': 1,
													'endRowIndex': 1000,
													'startColumnIndex': c,
													'endColumnIndex': c + 1,
												},
												'cell': {
													'userEnteredFormat': {
														'numberFormat': {
															'type': 'NUMBER',
															'pattern': '#,##0'
														}
													}
												},
												'fields': 'userEnteredFormat.numberFormat'
											}
										})
									if requests:
										sh.batch_update({'requests': requests})
						except Exception:
							LOG.debug('upload_to_google_sheets: fallo aplicando formato numérico en %s', safe_name, exc_info=True)
					except Exception:
						LOG.exception('upload_to_google_sheets: fallo preparando datos para %s', safe_name)
				LOG.info('upload_to_google_sheets: hoja %s actualizada (sheet fuente: %s, filas=%d)', safe_name, first_sheet, len(rows))
			except Exception:
				LOG.exception('upload_to_google_sheets: fallo al escribir target_sheet %s', target_sheet)
		else:
			# caso original: escribir cada hoja en su propia worksheet
			for sheet_name, rows in extracted.items():
				# sanitizar nombre de hoja
				safe_name = str(sheet_name)[:100]
				try:
					try:
						ws = sh.worksheet(safe_name)
					except Exception:
						# crear hoja si no existe
						ws = sh.add_worksheet(title=safe_name, rows=max(100, len(rows) + 5), cols=20)

					# Preserve existing header row if present; otherwise derive from extracted data.
					try:
						existing_headers = ws.row_values(1)
					except Exception:
						existing_headers = []

					if not existing_headers:
						headers = list(rows[0].keys()) if rows else []
						existing_headers = headers

					# Decide header for column A (fecha) and data headers for B..
					def _norm(s: str) -> str:
						return re.sub(r'\s+', ' ', str(s).strip().lower())

					header_a = 'fecha'
					data_headers = list(existing_headers)
					if data_headers:
						first_norm = _norm(data_headers[0])
						if 'fecha' in first_norm or 'date' in first_norm:
							header_a = existing_headers[0]
							data_headers = existing_headers[1:]

					# Clear sheet data and restore header (preserve row 1)
					date_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
					if clear:
						try:
							try:
								rc = getattr(ws, 'row_count', None)
								if rc and isinstance(rc, int) and rc > 1:
									ws.batch_clear([f"2:{rc}"])
								else:
									ws.batch_clear(["2:1000"])
							except Exception:
								ws.batch_clear(["2:1000"])
						except Exception:
							pass

					# Additionally, ensure columns D and E (from row 2 downward) are cleared
					# but only for the Sheet named 'Sheet1' to avoid touching other tabs.
					try:
						try:
							sn_local = str(safe_name).strip().lower() if safe_name else ''
							if sn_local == 'sheet1':
								rc_cols = getattr(ws, 'row_count', None) or 1000
								if isinstance(rc_cols, int) and rc_cols > 1:
									ws.batch_clear([f"D2:D{rc_cols}", f"E2:E{rc_cols}"])
								else:
									ws.batch_clear(["D2:D1000", "E2:E1000"])
						except Exception:
							# best-effort fallback when attempting to clear
							try:
								if sn_local == 'sheet1':
									ws.batch_clear(["D2:D1000", "E2:E1000"])
							except Exception:
								pass
					except Exception:
						pass

					try:
						# Write header row only if the sheet has no header yet and
						# do NOT touch row 1 for Sheet1 or Sheet2
						sn = str(safe_name).strip().lower() if safe_name else ''
						header_row = [header_a] + data_headers
						if not existing_headers and sn not in ('sheet1', 'sheet2'):
							try:
								ws.update('A1', [header_row])
							except Exception:
								LOG.debug('upload_to_google_sheets: fallo al escribir encabezado en %s', safe_name, exc_info=True)
					except Exception:
						LOG.debug('upload_to_google_sheets: fallo al preparar encabezado en %s', safe_name, exc_info=True)

					if not rows:
						LOG.info('upload_to_google_sheets: hoja %s actualizada (0 filas)', safe_name)
						continue

					# construir filas de datos (sin encabezado) usando el orden de data_headers
					try:
						extracted_keys = list(rows[0].keys()) if rows else []
						norm_map = { _norm(k): k for k in extracted_keys }

						# detect date key in extracted rows
						date_key = None
						for ek in extracted_keys:
							enk = _norm(ek)
							if 'fecha' in enk or 'date' in enk or 'dia' in enk:
								date_key = ek
								break

						mapped_keys = []
						for h in data_headers:
							nh = _norm(h)
							mapped = norm_map.get(nh)
							if not mapped:
								nh_simple = re.sub(r'[^0-9a-z]', '', nh)
								for ek_norm, ek in norm_map.items():
									if re.sub(r'[^0-9a-z]', '', ek_norm) == nh_simple:
										mapped = ek
										break
							if not mapped:
								for ek_norm, ek in norm_map.items():
									if nh in ek_norm or ek_norm in nh:
										mapped = ek
										break
							mapped_keys.append(mapped)

						table_data = []
						# choose sanitizer: only Sheet2 gets apostrophe-strip, Sheet1 and others are left untouched
						sn = str(safe_name).strip().lower() if safe_name else ''
						if sn == 'sheet2':
							base_sanitizer = _strip_leading_apostrophe
						else:
							base_sanitizer = _identity_sanitize

						# columns to coerce to numeric (sheet columns C,D,E,G,I,K) -> zero-based column indices
						numeric_col_letters = ['C', 'D', 'E', 'G', 'I', 'K']
						def _col_letter_to_index(letter: str) -> int:
							return ord(letter.upper()) - ord('A')
						numeric_col_indexes = {_col_letter_to_index(l) for l in numeric_col_letters}
						# data_headers correspond to sheet columns starting at B (col_index 1)
						# rowvals index -> sheet column index = index + 1
						numeric_rowvals_indexes = {ci - 1 for ci in numeric_col_indexes if ci >= 1}

						for r in rows:
							if date_key:
								pv = r.get(date_key, '')
								per_row_date = str(pv) if pv not in (None, '') else date_str
							else:
								per_row_date = date_str

							rowvals = []
							for idx, mk in enumerate(mapped_keys):
								if mk:
									raw = base_sanitizer(r.get(mk, ''))
									# Special-case: Sheet1 column C (data_headers start at B -> idx==1)
									if sn == 'sheet1' and idx == 1:
										v = _strip_dots_and_drop_decimals(raw)
										rowvals.append(v)
									elif sn == 'sheet2' and idx in numeric_rowvals_indexes:
										coerced = _sanitize_cell_value(raw)
										rowvals.append(coerced)
									else:
										rowvals.append(raw)
								else:
									rowvals.append('')
							table_data.append([per_row_date] + rowvals)

						# actualizar en bloque a partir de A2
						try:
							# Pre-write: ensure D2:E... are cleared to avoid ARRAYFORMULA or residual data (Sheet1 only)
							sn_per = str(safe_name).strip().lower() if safe_name else ''
							if sn_per == 'sheet1':
								try:
									try:
										rc_pre = getattr(ws, 'row_count', None) or (1 + len(table_data))
										if isinstance(rc_pre, int) and rc_pre > 1:
											ws.batch_clear([f"D2:D{rc_pre}", f"E2:E{rc_pre}"])
										else:
											ws.batch_clear(["D2:D1000", "E2:E1000"])
									except Exception:
										ws.batch_clear(["D2:D1000", "E2:E1000"])
								except Exception:
									pass

							ws.update('A2', table_data)
							# Post-write: asegurar limpieza específica en D2:E... tras escribir los datos (Sheet1 only)
							if sn_per == 'sheet1':
								try:
									try:
										rc_after = getattr(ws, 'row_count', None) or (1 + len(table_data))
										if isinstance(rc_after, int) and rc_after > 1:
											ws.batch_clear([f"D2:D{rc_after}", f"E2:E{rc_after}"])
										else:
											ws.batch_clear(["D2:D1000", "E2:E1000"])
									except Exception:
										ws.batch_clear(["D2:D1000", "E2:E1000"])
								except Exception:
									pass
							LOG.info('upload_to_google_sheets: hoja %s actualizada (%d filas)', safe_name, len(rows))
						except Exception:
							LOG.exception('upload_to_google_sheets: fallo al escribir datos en %s', safe_name)
					except Exception:
						LOG.exception('upload_to_google_sheets: fallo preparando datos para hoja %s', safe_name)
				except Exception:
					LOG.exception('upload_to_google_sheets: fallo al escribir hoja %s', sheet_name)
					continue

		return True
	except Exception:
		LOG.exception('upload_to_google_sheets: excepción inesperada')
		return False


# Opcional: si el usuario pone variables de entorno, llamar automáticamente tras la extracción.
# Estas variables NO se añaden aquí; el usuario debe proporcionar la ruta al JSON y el spreadsheet id.
# Ejemplo env vars esperadas: GOOGLE_SERVICE_ACCOUNT_JSON, GOOGLE_SHEET_ID
def _maybe_auto_upload(extracted: dict) -> None:
	try:
		# valores por defecto (proporcionados por el usuario). Preferir env vars si existen.
		default_sa = r'C:\Users\jperdomolc\Pictures\Qlik\estados-475119-24642bda896a.json'
		default_sid = '1LTiGfBQd_Qd6zhmCGEHpX0Jgaa3KuMkuuE8oHwQ6x3M'
		sa = _os.environ.get('GOOGLE_SERVICE_ACCOUNT_JSON', default_sa)
		sid = _os.environ.get('GOOGLE_SHEET_ID', default_sid)
		target = _os.environ.get('GOOGLE_SHEET_TAB', 'Sheet2')

		if sa and sid:
			LOG.info('Intentando subida automática a Google Sheets (target tab=%s)...', target)
			ok = upload_to_google_sheets(extracted, sid, sa, clear=True, target_sheet=target)
			if ok:
				LOG.info('Subida automática a Google Sheets finalizada con éxito')
			else:
				LOG.info('Subida automática a Google Sheets falló')
		else:
			LOG.debug('No hay credenciales/ID disponibles para Google Sheets')
	except Exception:
		LOG.debug('_maybe_auto_upload: fallo', exc_info=True)

def grid_listo(driver: webdriver.Chrome, selector: str, timeout: float = 20.0) -> bool:
    """Verificar si el grid está listo (visible y con contenido).
    
    Intenta encontrar el elemento y comprobar que está visible.
    Devuelve True si el grid está listo, False si no.
    """
    try:
        end = time.time() + float(timeout)
        while time.time() < end:
            try:
                el = driver.find_element(By.CSS_SELECTOR, selector)
                if el.is_displayed():
                    LOG.debug("grid_listo: elemento visible y listo: %s", selector)
                    return True
            except Exception:
                pass
            time.sleep(0.5)
        LOG.debug("grid_listo: timeout esperando elemento: %s", selector)
        return False
    except Exception:
        LOG.exception("grid_listo: excepción inesperada")
        return False


def run_once() -> None:
	url = (
		"https://qlik.copservir.com/sense/app/d39c40fb-a304-4eaf-9a30-50b7279d33f1/"
		"sheet/4f191cdb-aa40-409d-86b2-497a427a8b6a/state/analysis"
	)
	username = "Qlikzona29"
	password = "pF2A3f2x*"

	logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
	LOG.info('Starting minimal Qlik autofill (single run)')
	driver = setup_driver()
	submit_sent = False
	wrote_pwd = False
	try:
		LOG.info("Opening %s", url)
		driver.get(url)
		time.sleep(25)
		initial_process_done = False
		try:
			# Intentar traer al frente el navegador abierto por este script
			try:
				brought = bring_browser_to_front(driver)
			except Exception:
				LOG.debug('No se pudo forzar foco inicial en el navegador', exc_info=True)
				brought = False

			# Señalizar si el paso inicial tuvo éxito (bring_browser_to_front devolvió True)
			initial_process_done = bool(brought)
			LOG.info("Initial setup completed: %s", initial_process_done)
		except Exception:
			LOG.debug('Error during initial setup', exc_info=True)
			initial_process_done = False
			time.sleep(5)

		try:
			WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#grid > div:nth-child(8)")))
			LOG.info("Elemento '%s' está presente", "#grid > div:nth-child(8)")
			if focus_on_selector(driver, "#grid > div:nth-child(8)", timeout=3.0):
				LOG.info("Elemento '%s' enfocado correctamente", "#grid > div:nth-child(8)")
			else:
				LOG.info("No se pudo enfocar el selector '%s' (continuando)", "#grid > div:nth-child(8)")
		except Exception:
			LOG.debug('Error al esperar o enfocar el selector focus_... ', exc_info=True)
			time.sleep(2)

		try:
			# Asegurar foco antes de escribir el username
			try:
				bring_browser_to_front(driver)
			except Exception:
				LOG.debug('bring_browser_to_front falló antes de tipear username', exc_info=True)
		except Exception:
			pass
		ok = type_like_keyboard(driver, username, delay=0.08, click_first=True)

		success = False
		try:
			active = driver.switch_to.active_element
			val = active.get_attribute('value') or ''
			if username in val:
				success = True
		except Exception:
			pass

		try:
			parsed = urllib.parse.urlparse(url)
			host = parsed.netloc
		except Exception:
			host = 'qlik'

		if not success:
			LOG.info('Username no detectado; enviando por sistema')
			if send_text_via_system(driver, username, delay=0.08):
				LOG.info("Envío por sistema realizado para '%s'", username)
				try:
					if send_keys_via_pywinauto('{TAB}', host):
						LOG.info('Tab enviado vía pywinauto (tras username)')
				except Exception:
					pass
		else:
			LOG.info('Username detectado en elemento activo')

		LOG.info('Enviando password por sistema')
		if send_text_via_system(driver, password, delay=0.08):
			wrote_pwd = True
			LOG.info('Password enviado por sistema')
			try:
				if send_keys_via_pywinauto('{TAB}{ENTER}', host):
					submit_sent = True
					LOG.info('Submit intentado via pywinauto (final)')
				else:
					LOG.debug('pywinauto no envió submit, intentando Enter por keybd_event')
					_send_enter_windows()
			except Exception:
				LOG.debug('Error intentando submit via pywinauto', exc_info=True)

		try:
			# Aumentamos el tiempo de espera después del submit porque
			# la aplicación muestra una pantalla de carga que puede tardar.
			_keep_open_seconds = 30
			if submit_sent or wrote_pwd:
				LOG.info('Manteniendo navegador abierto %s segundos para inspección (post-login)...', _keep_open_seconds)
				time.sleep(_keep_open_seconds)
				
			   # --- Mostrar mes anterior ---
			try:
				# 1. Obtener fecha actual del sistema
				hoy = datetime.now()
				mes_actual = hoy.month
				year_actual = hoy.year
				
				# Calcular mes anterior (si es enero, el anterior es 12)
				mes_anterior = 12 if mes_actual == 1 else mes_actual - 1
				
				# Configuración de espera explícita
				wait = WebDriverWait(driver, 40)
				xpath_contenedor = (
					'//*[@id="qv-page-container"]/div[3]/div[1]/div/div[4]/div[2]/div[2]/div/'
					'div[4]/div[1]/div/div/div[1]/div/div[1]'
				)

				# --- PASO 1: Abrir contenedor ---
				btn_contenedor = wait.until(
					EC.element_to_be_clickable((By.XPATH, xpath_contenedor))
				)
				btn_contenedor.click()
				time.sleep(4) # Pausa solicitada tras abrir
				
				# --- PASO 2: Escribir mes anterior ---
				actions = webdriver.ActionChains(driver)
				actions.send_keys(str(mes_anterior)).perform()
				time.sleep(3) # Pausa para que el buscador de Qlik filtre
				
				# --- PASO 3: Navegar con teclado (Seleccionar el mes) ---
				actions.send_keys(Keys.TAB).send_keys(Keys.TAB).send_keys(Keys.SPACE).perform()
				time.sleep(4) # Pausa para procesar la selección
				
				# --- PASO 4: Aplicar primera vez ---
				btn_aplicar = wait.until(
					EC.element_to_be_clickable((By.XPATH, '//*[@id="actions-toolbar"]/div[4]/div[3]/button'))
				)
				btn_aplicar.click()
				time.sleep(2) # Pausa tras clic en aplicar

				# 🔒 Espera a que desaparezca el botón (indica que Qlik terminó de recalcular)
				wait.until(EC.invisibility_of_element(btn_aplicar))
				
				# --- PASO 5: Reabrir para quitar mes actual ---
				btn_contenedor = wait.until(
					EC.element_to_be_clickable((By.XPATH, xpath_contenedor))
				)
				btn_contenedor.click()
				time.sleep(3.5) # Pausa para que cargue la lista de selección
				
				# --- PASO 6: Quitar selección actual (Navegación teclado) ---
				actions_2 = webdriver.ActionChains(driver)
				actions_2.send_keys(Keys.TAB).send_keys(Keys.ARROW_DOWN).send_keys(Keys.SPACE).perform()
				time.sleep(3) # Pausa tras desmarcar el mes
				
				# --- PASO 7: Aplicar nuevamente ---
				btn_aplicar_final = wait.until(
					EC.element_to_be_clickable((By.XPATH, '//*[@id="actions-toolbar"]/div[4]/div[3]/button'))
				)
				btn_aplicar_final.click()
				
				# 🔒 Espera FINAL
				wait.until(EC.invisibility_of_element(btn_aplicar_final))
				time.sleep(4) # Pausa final de seguridad

				LOG.info("Proceso completado: Mes anterior (%s) seleccionado.", mes_anterior)

				# Define el selector del grid relevante UNA SOLA VEZ
				grid_sel = "#grid > div:nth-child(8)"
				LOG.info("Esperando grid después del cambio de mes: %s", grid_sel)
				
				# Espera SOLO UNA VEZ a que el grid esté visible
				WebDriverWait(driver, 30).until(
					EC.visibility_of_element_located((By.CSS_SELECTOR, grid_sel))
				)
				LOG.info("Grid visible después del cambio de mes: %s", grid_sel)
				
				# Comprueba si el grid está listo (usa SIEMPRE grid_sel)
				if not grid_listo(driver, grid_sel, timeout=20):
					LOG.warning("Grid no listo, se omite hover/export: %s", grid_sel)
					return
				
				# Trae el navegador al frente (opcional)
				try:
					bring_browser_to_front(driver)
				except Exception:
					LOG.debug('bring_browser_to_front falló antes del hover post-login', exc_info=True)
				
				# Hover sobre el grid (usa grid_sel)
				if hover_on_selector(driver, grid_sel, timeout=5.0):
					LOG.info("hover_on_selector: hover realizado correctamente en %s", grid_sel)
					try:
						# Dar tiempo suficiente para que el usuario vea el hover en pantalla
						time.sleep(15)
					except Exception:
						pass
					try:
						# Después del hover, localizar el botón "Más" y clickarlo
						btn_sel = (
							'#grid > div:nth-child(8) > '
							'div.object-and-panel-wrapper > div > '
							'div.ng-isolate-scope.detached-object-nav-wrapper > div '
							'button[tid="nav-menu-move"]'
						)
						if click_button_by_selector(driver, btn_sel, timeout=5.0):
							LOG.info("Botón 'Más' clicado correctamente: %s", btn_sel)
							try:
								time.sleep(10)
							except Exception:
								pass
							try:
								# Secuencia de menú: 'Descargar como...' -> 'Datos' -> 'Exportar'
								export_group_sel = '#export-group'
								export_sel = '#export'
								export_button = 'button[tid="table-export"]'

								if click_button_by_selector(driver, export_group_sel, timeout=5.0):
									LOG.info("click: export-group encontrado y clicado: %s", export_group_sel)
									try:
										time.sleep(0.6)
									except Exception:
										pass
									if click_button_by_selector(driver, export_sel, timeout=5.0):
										LOG.info("click: export encontrado y clicado: %s", export_sel)
										try:
											time.sleep(0.6)
										except Exception:
											pass
										if click_button_by_selector(driver, export_button, timeout=5.0):
											LOG.info("click: botón Exportar clicado: %s", export_button)
											try:
												time.sleep(1)
											except Exception:
												pass
											try:
												# Registrar tiempo de inicio de descarga y clicar el enlace de export
												download_start_ts = time.time()
												if click_export_url(driver, selector='a.export-url', timeout=10.0):
													LOG.info("click_export_url: enlace de descarga clicado correctamente")
													try:
														# Dar tiempo para que comience la descarga
														time.sleep(8)
													except Exception:
														pass

													# Intentar localizar el .xlsx descargado en Descargas
													downloads_dir = os.path.join(Path.home(), 'Downloads')
													found = find_latest_downloaded_file(downloads_dir, pattern='*.xlsx', since_ts=download_start_ts, timeout=30.0)
													if found:
														LOG.info('Archivo descargado detectado: %s', found)
														extracted = extract_excel_contents(found)
														if extracted is not None:
															out_file = Path('exported_data.json')
															try:
																with out_file.open('w', encoding='utf-8') as fh:
																	json.dump(extracted, fh, ensure_ascii=False, indent=2)
																LOG.info('Contenido del Excel guardado en %s', str(out_file))
															except Exception:
																LOG.exception('No se pudo escribir %s', str(out_file))
															try:
																# Intentar subida automática a Google Sheets si está configurado
																_maybe_auto_upload(extracted)
															except Exception:
																LOG.debug('Fallo al intentar subida automática a Google Sheets', exc_info=True)
															try:
																# Eliminar el fichero .xlsx descargado
																try:
																	p = Path(found)
																	if p.exists():
																		p.unlink()
																		LOG.info('Archivo descargado eliminado: %s', str(p))
																except Exception:
																	LOG.debug('No se pudo eliminar el archivo descargado %s', found, exc_info=True)
															except Exception:
																LOG.debug('Error al intentar remover archivo descargado', exc_info=True)
															# Después de eliminar el .xlsx, navegar al segundo link
															try:
																segunda_url = (
																	"https://qlik.copservir.com/sense/app/d39c40fb-a304-4eaf-9a30-50b7279d33f1/"
																	"sheet/28e2a154-adf5-4d68-9667-ee07b3bf9cf9/state/analysis"
																)
																LOG.info('Navegando a la segunda URL: %s', segunda_url)
																driver.get(segunda_url)
																time.sleep(30)
																try:
																	bring_browser_to_front(driver)
																except Exception:
																	LOG.debug('bring_browser_to_front falló en segunda URL', exc_info=True)

																
																# --- Iniciar interacción completa en la segunda URL ---
																try:
																	sel2 = '//*[@id="grid"]/div[17]'
																	try:
																		if hover_on_xpath(driver, sel2, timeout=5.0):
																			LOG.info('hover_on_xpath: hover realizado en segunda URL en %s', sel2)
																			try:
																				time.sleep(6)
																			except Exception:
																				pass

																			btn_sel2 = (
																				'#grid > div:nth-child(17) > '
																				'div.object-and-panel-wrapper > div > '
																				'div.ng-isolate-scope.detached-object-nav-wrapper > div '
																				'button[tid="nav-menu-move"]'
																			)
																			if click_button_by_selector(driver, btn_sel2, timeout=5.0):
																				LOG.info("Botón 'Más' clicado correctamente en segunda URL: %s", btn_sel2)
																				try:
																					time.sleep(1)
																				except Exception:
																					pass
																				export_group_sel2 = '#export-group'
																				export_sel2 = '#export'

																				if click_button_by_selector(driver, export_group_sel2, timeout=5.0):
																					LOG.info("click: export-group encontrado y clicado en segunda URL: %s", export_group_sel2)
																					try:
																						time.sleep(0.6)
																					except Exception:
																						pass
																					if click_button_by_selector(driver, export_sel2, timeout=5.0):
																						LOG.info("click: export encontrado y clicado en segunda URL: %s", export_sel2)
																						try:
																							time.sleep(0.6)
																						except Exception:
																							pass

																						# En esta segunda URL no existe el botón 'table-export'.
																						# Directamente intentar clicar el enlace de descarga (a.export-url)
																						try:
																							download_start_ts2 = time.time()
																							# Primero intentar el anchor conocido
																							if click_export_url(driver, selector='a.export-url', timeout=6.0):
																								LOG.info("click_export_url: enlace de descarga clicado correctamente en segunda URL")
																								clicked_download = True
																							else:
																								# Fallback: buscar anchors con .xlsx o texto 'export'/'exportar'
																								if click_export_link_with_fallback(driver, timeout=6.0):
																									LOG.info("click_export_link_with_fallback: enlace de descarga clicado en segunda URL")
																									clicked_download = True
																								else:
																									clicked_download = False

																							if clicked_download:
																								try:
																									time.sleep(8)
																								except Exception:
																									pass

																								downloads_dir2 = os.path.join(Path.home(), 'Downloads')
																								found2 = find_latest_downloaded_file(downloads_dir2, pattern='*.xlsx', since_ts=download_start_ts2, timeout=30.0)
																								if found2:
																									LOG.info('Archivo descargado detectado en segunda URL: %s', found2)
																									extracted2 = extract_excel_contents(found2)
																									if extracted2 is not None:
																										out_file2 = Path('exported_data_2.json')
																										try:
																											with out_file2.open('w', encoding='utf-8') as fh:
																												json.dump(extracted2, fh, ensure_ascii=False, indent=2)
																											LOG.info('Contenido del Excel guardado en %s', str(out_file2))
																										except Exception:
																											LOG.exception('No se pudo escribir %s', str(out_file2))

																										# Subida automática con target tab por defecto en 'Sheet1'
																										try:
																											default_sa = r'C:\Users\jperdomolc\Pictures\Qlik\estados-475119-24642bda896a.json'
																											default_sid = '1LTiGfBQd_Qd6zhmCGEHpX0Jgaa3KuMkuuE8oHwQ6x3M'
																											sa = _os.environ.get('GOOGLE_SERVICE_ACCOUNT_JSON', default_sa)
																											sid = _os.environ.get('GOOGLE_SHEET_ID', default_sid)
																											target = _os.environ.get('GOOGLE_SHEET_TAB', 'Sheet1')

																											if sa and sid:
																												LOG.info('Intentando subida automática a Google Sheets (target tab=%s)...', target)
																												ok = upload_to_google_sheets(extracted2, sid, sa, clear=True, target_sheet=target)
																												if ok:
																													LOG.info('Subida automática a Google Sheets (Sheet1) finalizada con éxito')
																												else:
																													LOG.info('Subida automática a Google Sheets (Sheet1) falló')
																											else:
																												LOG.debug('No hay credenciales/ID disponibles para Google Sheets (segunda URL)')
																										except Exception:
																											LOG.debug('Fallo al intentar subida automática a Google Sheets en segunda URL', exc_info=True)
																										try:
																											p2 = Path(found2)
																											if p2.exists():
																												p2.unlink()
																												LOG.info('Archivo descargado eliminado en segunda URL: %s', str(p2))
																										except Exception:
																											LOG.debug('No se pudo eliminar el archivo descargado en segunda URL %s', found2, exc_info=True)
																									else:
																										LOG.info('No se pudo extraer contenido del Excel en segunda URL: %s', found2)
																								else:
																									LOG.info('No se detectó archivo .xlsx en %s dentro del timeout en segunda URL', downloads_dir2)
																						except Exception:
																							LOG.debug('Error al intentar click_export_url o procesar descarga en segunda URL', exc_info=True)
																					else:
																						LOG.info('No se pudo clicar el item export (%s) en segunda URL', export_sel2)
																				else:
																					LOG.info('No se pudo clicar el grupo export-group (%s) en segunda URL', export_group_sel2)
																			else:
																				LOG.info("No se pudo clicar el botón 'Más' (%s) en segunda URL", btn_sel2)
																		else:
																			LOG.info('hover_on_xpath: no se pudo hacer hover en %s en segunda URL (continuando)', sel2)
																	except Exception:
																		LOG.debug('Error en la interacción de la segunda URL', exc_info=True)
																except Exception:
																	LOG.debug('Fallo preparando interacción en la segunda URL', exc_info=True)
															except Exception:
																LOG.debug('Error navegando a la segunda URL', exc_info=True)
														else:
															LOG.info('No se pudo extraer contenido del Excel: %s', found)
													else:
														LOG.info('No se detectó archivo .xlsx en %s dentro del timeout', downloads_dir)
												else:
													LOG.info("click_export_url: no se encontró el enlace de descarga (a.export-url)")
											except Exception:
												LOG.debug('Error al intentar click_export_url o procesar descarga', exc_info=True)
										else:
											LOG.info("No se pudo clicar el botón Exportar (%s)", export_button)
									else:
										LOG.info("No se pudo clicar el item export (%s)", export_sel)
								else:
									LOG.info("No se pudo clicar el grupo export-group (%s)", export_group_sel)
							except Exception:
								LOG.debug('Error al intentar clicar el botón Más', exc_info=True)
						else:
							LOG.info("No se pudo clicar el botón 'Más' (%s)", btn_sel)
					except Exception:
						LOG.debug('Error al intentar clicar el botón Más', exc_info=True)
				else:
					LOG.info("hover_on_selector: no se pudo hacer hover en %s (continuando)", grid_sel)
			except Exception:
				LOG.debug('Error en el flujo post-cambio de mes', exc_info=True)
		except Exception:
			pass
	finally:
		driver.quit()


def main() -> None:
	"""Loop runner: ejecuta `run_once()` inmediatamente y luego espera hasta las 06:00 local siguiente para repetir.

	Ctrl+C detiene el loop.
	"""
	try:
		while True:
			try:
				run_once()
			except Exception:
				LOG.exception('run_once: excepción no controlada durante la ejecución')

			# calcular próxima ejecución a las 06:00, 06:30 o 12:30 local (12:30 solo fines de semana)
			now = datetime.now()
			next_run_6_exact = now.replace(hour=6, minute=0, second=0, microsecond=0)
			next_run_6 = now.replace(hour=6, minute=30, second=0, microsecond=0)
			next_run_12 = now.replace(hour=12, minute=30, second=0, microsecond=0)
			
			# encontrar la próxima hora de ejecución
			candidates = [next_run_6_exact, next_run_6]
			if now.weekday() in [5, 6]:  # 5=Saturday, 6=Sunday
				candidates.append(next_run_12)
			future_candidates = [t for t in candidates if t > now]
			if future_candidates:
				next_run = min(future_candidates)
			else:
				# todas pasaron hoy, tomar la primera de mañana
				next_run = min(candidates) + timedelta(days=1)
			
			wait_seconds = (next_run - now).total_seconds()
			LOG.info('Siguiente ejecución programada para %s (en %d segundos)', next_run.isoformat(), int(wait_seconds))

			try:
				# dormir hasta la próxima ejecución (permitir interrumpir con Ctrl+C)
				time.sleep(wait_seconds)
			except KeyboardInterrupt:
				LOG.info('Interrupción recibida durante la espera; terminando.')
				break
	except KeyboardInterrupt:
		LOG.info('Interrupción recibida; saliendo')


if __name__ == '__main__':
	main()